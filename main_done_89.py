import openpyxl
from openpyxl import load_workbook as lwb
import warnings
from datetime import datetime as dt

print("Время запуска: ", dt.now())
warnings.simplefilter("ignore")

wb_result = lwb("результат.xlsx")
wb_source = lwb("PeriodChessCross3_1.xlsx")
ws_result = wb_result.active
ws_source = wb_source.active

def read_date(str_start, str_end, col_start, col_end):
    # Чтение части таблицы файла
    y_list = []
    result_list_2019 = []
    for i in range(str_start, (str_end + 1)):  # Значение в файле источника есть с 7 по 42 строку, с 4 по ... столбец
        for j in range(col_start, (col_end + 1)):
            y = str(ws_source.cell(i, j).value)
            if y != "None":  # y_list - это одна строка
                y_list.append(y)
        result_list_2019.append(y_list)  # Итоговый файл начинается не с 7 , а с 5 строки , поэтому добавляем смещение
        y_list = []
    # print("Сформированная матрица")
    # for i in range(0,len(result_list_2019)):
    #   print(result_list_2019[i])  #   !!! Раскоментить при просмотре таблицы считываемой !!!
    return result_list_2019


def write_date(use_list, str_start, str_end, col_str, col_end):
    # Запись в результирующий файл части таблицы
    for stroc in range(str_start, (str_end + 1)):  # Строка у нас начинается с 5 и заканчивается 40 , столбец с 4 и заканчивается 16
        for col in range(col_str, (col_end)):
            global ws_result
            ws_result.cell(stroc, col).value = float(use_list[stroc - str_start][col - col_str])
    # print("Запись в промежуточную таблицу проведена")

# строка начислени
write_date(read_date(7, 42, 4, 5), 5, 40, 4, 5)
# 2019-2019
write_date(read_date(7, 42, 5, 17), 5, 40, 5, 17)
# 2019-2020
write_date(read_date(7, 42, 19, 32), 5, 40, 17, 29)
# 2019-2021
write_date(read_date(7, 42, 34, 47), 5, 40, 29, 41)
# 2019-2022
write_date(read_date(7, 42, 48, 52), 5, 40, 41, 45)  # здесь добавлять следующий месяц 53 и 46
# 2020
# строка начислени
write_date(read_date(44, 79, 4, 5), 44, 79, 4, 5)
# 2020-2019
write_date(read_date(44, 79, 5, 17), 44, 79, 5, 17)
# 2020-2020
write_date(read_date(44, 79, 19, 32), 44, 79, 17, 29)
# 2020-2021
write_date(read_date(44, 79, 34, 47), 44, 79, 29, 41)
# 2020-2022
write_date(read_date(44, 79, 48, 52), 44, 79, 41, 45)  # здесь добавлять следующий месяц 53 и 46

# 2021
# строка начислени
write_date(read_date(81, 116, 4, 5), 83, 118, 4, 5)
# 2021-2019
write_date(read_date(81, 116, 5, 17), 83, 118, 5, 17)
# 2021-2020
write_date(read_date(81, 116, 19, 32), 83, 118, 17, 29)
# 2021-2021
write_date(read_date(81, 116, 34, 47), 83, 118, 29, 41)
# 2021-2022
write_date(read_date(81, 116, 48, 52), 83, 118, 41, 45)  # здесь добавлять следующий месяц 53 и 46

# 2022
# строка начислени
write_date(read_date(118, 125, 4, 5), 122, 129, 4, 5)
# 2022-2019
write_date(read_date(118, 129, 5, 17), 122, 133, 5, 17)
# 2022-2020
write_date(read_date(118, 129, 19, 32), 122, 133, 17, 29)
# 2022-2021
write_date(read_date(118, 129, 34, 47), 122, 133, 29, 41)
# 2022-2022
write_date(read_date(118, 129, 48, 52), 122, 133, 41, 45)  # здесь добавлять следующий месяц 53 и 46

data_today = dt.today()

name_f = f"89 на {data_today.strftime('%d')} {data_today.strftime('%m')} {data_today.strftime('%Y')}.xlsx"

wb_result.save(name_f)

time_now = dt.now()
print("Время завершения: ", time_now)
